# -*- coding: utf-8 -*-
"""
将 Quest 节点 Markdown 表格转换为 Excel
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_excel():
    # 创建工作簿
    wb = openpyxl.Workbook()

    # 删除默认工作表
    wb.remove(wb.active)

    # 定义样式
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

    category_font = Font(name='Microsoft YaHei', size=12, bold=True, color='FFFFFF')
    category_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    subcategory_font = Font(name='Microsoft YaHei', size=10, bold=True)
    subcategory_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    normal_font = Font(name='Microsoft YaHei', size=10)

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # ==================== 1. 核心节点 ====================
    ws_core = wb.create_sheet('核心节点')
    ws_core.append(['节点类型', '功能描述', '所在文件'])

    core_nodes = [
        ['NodeDefinition', '所有Quest节点的基类', 'node.h'],
        ['DisableableNodeDefinition', '可禁用的节点基类', 'disableableNode.h'],
        ['SignalStoppingNodeDefinition', '可阻断信号传播的节点基类', 'signalStoppingNode.h'],
        ['TypedSignalStoppingNodeDefinition', '带类型的信号阻断节点', 'signalStoppingNode.h'],
        ['StartEndNodeDefinition', '开始/结束节点基类', 'startEndNode.h'],
        ['StartNodeDefinition', 'Quest开始节点', 'startNode.h'],
        ['EndNodeDefinition', 'Quest结束节点', 'endNode.h'],
        ['IONodeDefinition', '输入/输出节点基类', 'ioNode.h'],
        ['InputNodeDefinition', '输入节点', 'inputNode.h'],
        ['OutputNodeDefinition', '输出节点', 'outputNode.h'],
        ['GraphDefinition', 'Quest图定义', 'graph.h'],
        ['SocketDefinition', 'Socket定义', 'socket.h'],
    ]

    for row in core_nodes:
        ws_core.append(row)

    format_sheet(ws_core, header_font, header_fill, normal_font, border, alignment_center, alignment_left)

    # ==================== 2. 角色管理器 ====================
    ws_char = wb.create_sheet('角色管理器')
    ws_char.append(['节点类型', '功能描述', '所在文件'])

    char_nodes = [
        ['CharacterManagerNodeDefinition', '角色管理器主节点', 'characterManagerNode.h'],
        ['CharacterManagerParameters_SetAttitudeGroupForPuppet', '设置AI态度组', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerParameters_SetGroupsAttitude', '设置组态度', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerParameters_SetMortality', '设置生死状态', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerParameters_SetAnimset', '设置动画集', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerParameters_SetLowGravity', '设置低重力', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerParameters_EnableBumps', '启用碰撞', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerParameters_SetStatusEffect', '设置状态效果', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerParameters_SetReactionPreset', '设置反应预设', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerParameters_SetGender', '设置性别', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerParameters_SetAsCrowdObstacle', '设为人群障碍物', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerParameters_SetProgressionBuild', '设置进度构建', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerParameters_SetLifePath', '设置人生轨迹', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerParameters_HealPlayer', '治疗玩家', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerCombat_ModifyHealth', '修改生命值', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerCombat_Kill', '杀死角色', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerCombat_EquipWeapon', '装备武器', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerCombat_SetWeaponState', '设置武器状态', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerCombat_SetDeathDirection', '设置死亡方向', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerCombat_ChangeLevel', '改变等级', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerCombat_ManageRagdoll', '管理布娃娃系统', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerCombat_AssignSquad', '分配小队', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerParameters_SetCombatSpace', '设置战斗空间', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerVisuals_ChangeEntityAppearance', '改变实体外观', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerVisuals_PrefetchEntityAppearance', '预加载实体外观', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerVisuals_GenitalsManager', '生殖器管理', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerVisuals_BreastSizeController', '胸部大小控制', 'characterManagerNodeTypeImpl.h'],
        ['CharacterManagerVisuals_SetBrokenNoseStage', '设置鼻梁破损阶段', 'characterManagerNodeTypeImpl.h'],
    ]

    for row in char_nodes:
        ws_char.append(row)

    format_sheet(ws_char, header_font, header_fill, normal_font, border, alignment_center, alignment_left)

    # ==================== 3. 实体管理器 ====================
    ws_entity = wb.create_sheet('实体管理器')
    ws_entity.append(['节点类型', '功能描述', '所在文件'])

    entity_nodes = [
        ['EntityManagerNodeDefinition', '实体管理器主节点', 'entityManagerNode.h'],
        ['EntityManagerSetAttachment_NodeType', '设置附着', 'entityManagerNodeTypeImpl.h'],
        ['EntityManagerSetDestructionState_NodeType', '设置破坏状态', 'entityManagerNodeTypeImpl.h'],
        ['EntityManagerManageBinkComponent_NodeType', '管理Bink组件', 'entityManagerNodeTypeImpl.h'],
        ['EntityManagerSetMeshAppearance_NodeType', '设置网格外观', 'entityManagerNodeTypeImpl.h'],
        ['EntityManagerEnablePlayerTPPRepresentation_NodeType', '启用玩家第三人称表示', 'entityManagerNodeTypeImpl.h'],
        ['EntityManagerToggleComponent_NodeType', '切换组件', 'entityManagerNodeTypeImpl.h'],
        ['EntityManagerChangeAppearance_NodeType', '改变外观', 'entityManagerNodeTypeImpl.h'],
        ['EntityManagerMountPuppet_NodeType', '骑乘Puppet', 'entityManagerNodeTypeImpl.h'],
        ['EntityManagerSendAnimationEvent_NodeType', '发送动画事件', 'entityManagerNodeTypeImpl.h'],
        ['EntityManagerSetStat_NodeType', '设置属性', 'entityManagerNodeTypeImpl.h'],
        ['EntityManagerToggleMirrorsArea_NodeType', '切换镜像区域', 'entityManagerNodeTypeImpl.h'],
        ['EntityManagerSetAttachment_ToActor', '附着到角色', 'entityManagerNodeTypeImpl.h'],
        ['EntityManagerDestroyCarriedObject', '销毁携带物体', 'entityManagerNodeTypeImpl.h'],
        ['EntityManagerSetAttachment_ToNode', '附着到节点', 'entityManagerNodeTypeImpl.h'],
        ['EntityManagerSetAttachment_ToWorld', '附着到世界', 'entityManagerNodeTypeImpl.h'],
    ]

    for row in entity_nodes:
        ws_entity.append(row)

    format_sheet(ws_entity, header_font, header_fill, normal_font, border, alignment_center, alignment_left)

    # ==================== 4. UI管理器 ====================
    ws_ui = wb.create_sheet('UI管理器')
    ws_ui.append(['节点类型', '功能描述', '所在文件'])

    ui_nodes = [
        ['UIManagerNodeDefinition', 'UI管理器主节点', 'uiManagerNode.h'],
        ['AddCombatLogMessage_NodeType', '添加战斗日志消息', 'uiManagerNodeType.h'],
        ['SwitchNameplate_NodeType', '切换名牌', 'uiManagerNodeType.h'],
        ['AddBraindanceClue_NodeType', '添加脑舞线索', 'uiManagerNodeType.h'],
        ['DiscoverBraindanceClue_NodeType', '发现脑舞线索', 'uiManagerNodeType.h'],
        ['DisplayMessageBox_NodeType', '显示消息框', 'uiManagerNodeType.h'],
        ['ProgressBar_NodeType', '进度条', 'uiManagerNodeType.h'],
        ['ProximityProgressBar_NodeType', '接近度进度条', 'uiManagerNodeType.h'],
        ['ShowDialogIndicator_NodeType', '显示对话指示器', 'uiManagerNodeType.h'],
        ['HUDVideo_NodeType', 'HUD视频', 'uiManagerNodeType.h'],
        ['SetLocationName_NodeType', '设置位置名称', 'uiManagerNodeType.h'],
        ['WarningMessage_NodeType', '警告消息', 'uiManagerNodeType.h'],
        ['ShowOnscreen_NodeType', '屏幕显示', 'uiManagerNodeType.h'],
        ['OverrideLoadingScreen_NodeType', '覆盖加载屏幕', 'uiManagerNodeType.h'],
        ['GlitchLoadingScreen_NodeType', '故障加载屏幕', 'uiManagerNodeType.h'],
        ['WaitForAnyKeyLoadingScreen_NodeType', '等待任意键加载屏幕', 'uiManagerNodeType.h'],
        ['SetUIGameContext_NodeType', '设置UI游戏上下文', 'uiManagerNodeType.h'],
        ['SetHUDEntryForcedVisibility_NodeType', '设置HUD条目强制可见性', 'uiManagerNodeType.h'],
        ['QuickItemsManager_NodeType', '快速物品管理器', 'uiManagerNodeType.h'],
        ['VendorPanel_NodeType', '商贩面板', 'uiManagerNodeType.h'],
        ['OpenBriefing_NodeType', '打开简报', 'uiManagerNodeType.h'],
        ['EnableBraindanceFinish_NodeType', '启用脑舞完成', 'uiManagerNodeType.h'],
        ['SwitchToScenario_NodeType', '切换到场景', 'uiManagerNodeType.h'],
        ['SetBriefingSize_NodeType', '设置简报大小', 'uiManagerNodeType.h'],
        ['SetBriefingAlignment_NodeType', '设置简报对齐', 'uiManagerNodeType.h'],
        ['ShowNarrativeEvent_NodeType', '显示叙事事件', 'uiManagerNodeType.h'],
        ['ShowCustomTooltip_NodeType', '显示自定义提示', 'uiManagerNodeType.h'],
        ['Tutorial_NodeType', '教程', 'uiManagerNodeType.h'],
        ['ToggleMinimapVisibility_NodeSubType', '切换小地图可见性', 'uiManagerNodeType.h'],
        ['ToggleStealthMappinVisibility_NodeSubType', '切换潜行地图标记可见性', 'uiManagerNodeType.h'],
        ['ShowHighlight_NodeSubType', '显示高亮', 'uiManagerNodeType.h'],
        ['ShowBracket_NodeSubType', '显示括号', 'uiManagerNodeType.h'],
        ['ShowOverlay_NodeSubType', '显示覆盖层', 'uiManagerNodeType.h'],
        ['ShowPopup_NodeSubType', '显示弹出窗口', 'uiManagerNodeType.h'],
        ['BriefingSequencePlayer_NodeType', '简报序列播放器', 'uiManagerNodeType.h'],
        ['TriggerIconGeneration_NodeType', '触发图标生成', 'uiManagerNodeType.h'],
        ['InputHint_NodeType', '输入提示', 'uiManagerNodeType.h'],
        ['InputHintGroup_NodeType', '输入提示组', 'uiManagerNodeType.h'],
        ['ShowLevelUpNotification_NodeType', '显示升级通知', 'uiManagerNodeType.h'],
        ['ShowCustomQuestNotification_NodeType', '显示自定义任务通知', 'uiManagerNodeType.h'],
        ['SetMetaQuestProgress_NodeType', '设置元任务进度', 'uiManagerNodeType.h'],
        ['SetSaveDataLoadingScreen_NodeType', '设置存档数据加载屏幕', 'uiManagerNodeType.h'],
        ['SetFastTravelBinksGroup_NodeType', '设置快速旅行视频组', 'uiManagerNodeType.h'],
        ['OpenPhotoMode_NodeType', '打开照片模式', 'uiManagerNodeType.h'],
        ['ShowPointOfNoReturnPrompt_NodeType', '显示不归路提示', 'uiManagerNodeType.h'],
        ['FinalBoardsVideosFinished_NodeType', '最终板视频完成', 'uiManagerNodeType.h'],
        ['FinalBoardsEnableSkipCredits_NodeType', '最终板启用跳过制作人员名单', 'uiManagerNodeType.h'],
        ['FinalBoardsOpenSpeakerScreen_NodeType', '最终板打开扬声器屏幕', 'uiManagerNodeType.h'],
    ]

    for row in ui_nodes:
        ws_ui.append(row)

    format_sheet(ws_ui, header_font, header_fill, normal_font, border, alignment_center, alignment_left)

    # ==================== 5. 车辆管理器 ====================
    ws_vehicle = wb.create_sheet('车辆管理器')
    ws_vehicle.append(['节点类型', '功能描述', '所在文件'])

    vehicle_nodes = [
        ['VehicleNodeDefinition', '车辆管理器主节点', 'vehicleManagerNode.h'],
        ['AssignCharacter_NodeType', '分配角色', 'vehicleManagerNodeType.h'],
        ['RequestVehicleCameraPerspective_NodeType', '请求车辆相机视角', 'vehicleManagerNodeType.h'],
        ['MoveOnSpline_NodeType', '在样条上移动', 'vehicleManagerNodeType.h'],
        ['ToggleCombatForPlayer_NodeType', '切换玩家战斗', 'vehicleManagerNodeType.h'],
        ['ToggleSwitchSeatsForPlayer_NodeType', '切换玩家座位', 'vehicleManagerNodeType.h'],
        ['MoveOnSplineAndKeepDistance_NodeType', '在样条上移动并保持距离', 'vehicleManagerNodeType.h'],
        ['MoveOnSplineControlRubberbanding_NodeType', '在样条上移动控制橡皮筋效果', 'vehicleManagerNodeType.h'],
        ['StartVehicle_NodeType', '启动车辆', 'vehicleManagerNodeType.h'],
        ['StopVehicle_NodeType', '停止车辆', 'vehicleManagerNodeType.h'],
        ['FollowObject_NodeType', '跟随物体', 'vehicleManagerNodeType.h'],
        ['ResetMovement_NodeType', '重置移动', 'vehicleManagerNodeType.h'],
        ['SetAutopilot_NodeType', '设置自动驾驶', 'vehicleManagerNodeType.h'],
        ['ToggleBrokenTire_NodeType', '切换轮胎损坏', 'vehicleManagerNodeType.h'],
        ['ToggleForceBrake_NodeType', '切换强制制动', 'vehicleManagerNodeType.h'],
        ['FlushAutopilot_NodeType', '刷新自动驾驶', 'vehicleManagerNodeType.h'],
        ['ToggleTankCustomFPPLockOff_NodeType', '切换坦克自定义FPP锁定', 'vehicleManagerNodeType.h'],
        ['ToggleWeaponEnabled_NodeType', '切换武器启用', 'vehicleManagerNodeType.h'],
        ['OverrideSplineSpeed_NodeType', '覆盖样条速度', 'vehicleManagerNodeType.h'],
        ['Repair_NodeType', '修理', 'vehicleManagerNodeType.h'],
        ['ToggleDoor_NodeType', '切换车门', 'vehicleManagerNodeType.h'],
        ['SpawnPlayerVehicle_NodeType', '生成玩家车辆', 'vehicleManagerNodeType.h'],
        ['Teleport_NodeType', '传送', 'vehicleManagerNodeType.h'],
        ['ForbiddenTrigger_NodeType', '禁止触发器', 'vehicleManagerNodeType.h'],
        ['EnableVehicleSummon_NodeType', '启用车辆召唤', 'vehicleManagerNodeType.h'],
        ['EnablePlayerVehicle_NodeType', '启用玩家车辆', 'vehicleManagerNodeType.h'],
        ['ToggleWindow_NodeType', '切换车窗', 'vehicleManagerNodeType.h'],
        ['UnassignAll_NodeType', '取消所有分配', 'vehicleManagerNodeType.h'],
        ['ForcePhysicsWakeUp_NodeType', '强制物理唤醒', 'vehicleManagerNodeType.h'],
        ['SetImmovable_NodeType', '设置不可移动', 'vehicleManagerNodeType.h'],
    ]

    for row in vehicle_nodes:
        ws_vehicle.append(row)

    format_sheet(ws_vehicle, header_font, header_fill, normal_font, border, alignment_center, alignment_left)

    # ==================== 6. AI命令节点 ====================
    ws_ai = wb.create_sheet('AI命令节点')
    ws_ai.append(['节点类型', '功能描述', '所在文件'])

    ai_nodes = [
        ['AICommandNodeBase', 'AI命令节点基类', 'aiCommandNodeBase.h'],
        ['ConfigurableAICommandNode', '可配置AI命令节点', 'questConfigurableAICommandNode.h'],
        ['SendAICommandNodeDefinition', '发送AI命令', 'sendAICommandNode.h'],
        ['CombatNodeDefinition', '战斗节点', 'combatNode.h'],
        ['MovePuppetNodeDefinition', '移动Puppet', 'movePuppetNode.h'],
        ['MiscAICommandNode', '杂项AI命令', 'miscAICommandNode.h'],
        ['TeleportPuppetNodeDefinition', '传送Puppet', 'teleportPuppetNode.h'],
        ['EquipItemNodeDefinition', '装备物品', 'equipItemNode.h'],
        ['UnequipItemNodeDefinition', '卸下物品', 'unequipItemNode.h'],
        ['UseWorkspotNodeDefinition', '使用工作点', 'useWorkspotNode.h'],
        ['RotateToNodeDefinition', '旋转到目标', 'rotateToNode.h'],
        ['VehicleNodeCommandDefinition', '车辆命令', 'vehicleCommandNode.h'],
        ['ForcedBehaviourNodeDefinition', '强制行为', 'forcedBehaviourNode.h'],
        ['ClearForcedBehavioursNodeDefinition', '清除强制行为', 'forcedBehaviourNode.h'],
        ['LookAtDrivenTurnsNode', '注视驱动转向', 'lookAtDrivenTurnsNode.h'],
    ]

    for row in ai_nodes:
        ws_ai.append(row)

    format_sheet(ws_ai, header_font, header_fill, normal_font, border, alignment_center, alignment_left)

    # ==================== 7. 逻辑节点 ====================
    ws_logic = wb.create_sheet('逻辑节点')
    ws_logic.append(['节点类型', '功能描述', '所在文件'])

    logic_nodes = [
        ['LogicalBaseNodeDefinition', '逻辑节点基类', 'logicalBaseNode.h'],
        ['LogicalAndNodeDefinition', '逻辑与节点', 'logicalAndNode.h'],
        ['LogicalXorNodeDefinition', '逻辑异或节点', 'logicalXorNode.h'],
        ['LogicalHubNodeDefinition', '逻辑Hub节点', 'logicalHubNode.h'],
    ]

    for row in logic_nodes:
        ws_logic.append(row)

    format_sheet(ws_logic, header_font, header_fill, normal_font, border, alignment_center, alignment_left)

    # ==================== 8. 条件节点 ====================
    ws_condition = wb.create_sheet('条件节点')
    ws_condition.append(['节点类型', '功能描述', '所在文件'])

    condition_nodes = [
        # 基础类
        ['IBaseCondition', '条件基类接口', 'condition.h'],
        ['Condition', '条件类', 'condition.h'],
        ['TypedCondition', '带类型的条件', 'condition.h'],
        ['LogicalCondition', '逻辑条件', 'logicalCondition.h'],
        ['ConditionNodeDefinition', '条件节点', 'conditionNode.h'],
        ['PauseConditionNodeDefinition', '暂停条件节点', 'pauseConditionNode.h'],
        # 对象条件
        ['ObjectCondition', '对象条件', 'objectCondition.h'],
        ['Interaction_ConditionType', '交互条件', 'objectConditionType.h'],
        ['Inventory_ConditionType', '库存条件', 'objectConditionType.h'],
        ['Inspect_ConditionType', '检查条件', 'objectConditionType.h'],
        ['Scan_ConditionType', '扫描条件', 'objectConditionType.h'],
        ['EntryScanned_ConditionType', '条目扫描条件', 'objectConditionType.h'],
        ['Device_ConditionType', '设备条件', 'objectConditionType.h'],
        ['Destruction_ConditionType', '破坏条件', 'objectConditionType.h'],
        ['Tagged_ConditionType', '标记条件', 'objectConditionType.h'],
        # 支付条件
        ['PaymentCondition', '支付条件', 'paymentCondition.h'],
        ['PaymentBalanced_ConditionType', '平衡支付条件', 'paymentConditionType.h'],
        ['PaymentFixedAmount_ConditionType', '固定金额支付条件', 'paymentConditionType.h'],
        # 属性条件
        ['StatsCondition', '属性条件', 'statsCondition.h'],
        ['Stat_ConditionType', '属性条件', 'statsConditionType.h'],
        ['StreetCredTier_ConditionType', '街头声望等级条件', 'statsConditionType.h'],
        ['LifePath_ConditionType', '人生轨迹条件', 'statsConditionType.h'],
        ['Build_ConditionType', '构建条件', 'statsConditionType.h'],
        # 系统条件
        ['CameraFocus_ConditionType', '相机焦点条件', 'systemConditionType.h'],
        ['VisionMode_ConditionType', '视觉模式条件', 'systemConditionType.h'],
        ['Platform_ConditionType', '平台条件', 'systemConditionType.h'],
        ['InputAction_ConditionType', '输入动作条件', 'systemConditionType.h'],
        ['InputController_ConditionType', '输入控制器条件', 'systemConditionType.h'],
        ['Phone_ConditionType', '电话条件', 'systemConditionType.h'],
        ['PhonePickUp_ConditionType', '电话接听条件', 'systemConditionType.h'],
        ['Prereq_ConditionType', '前置条件', 'systemConditionType.h'],
        ['Weather_ConditionType', '天气条件', 'systemConditionType.h'],
        ['Radio_ConditionType', '电台条件', 'systemConditionType.h'],
        ['RadioTrack_ConditionType', '电台曲目条件', 'systemConditionType.h'],
        ['PlaylistTrackChanged_ConditionType', '播放列表曲目变更条件', 'systemConditionType.h'],
        ['Language_ConditionType', '语言条件', 'systemConditionType.h'],
        ['GOGReward_ConditionType', 'GOG奖励条件', 'systemConditionType.h'],
        ['SaveLock_ConditionType', '存档锁定条件', 'systemConditionType.h'],
        # 时间条件
        ['TimeCondition', '时间条件', 'timeCondition.h'],
        ['RealtimeDelay_ConditionType', '实时延迟条件', 'timeConditionType.h'],
        ['GameTimeDelay_ConditionType', '游戏时间延迟条件', 'timeConditionType.h'],
        ['TimePeriod_ConditionType', '时间段条件', 'timeConditionType.h'],
    ]

    for row in condition_nodes:
        ws_condition.append(row)

    format_sheet(ws_condition, header_font, header_fill, normal_font, border, alignment_center, alignment_left)

    # ==================== 9. 其他管理器 ====================
    ws_other = wb.create_sheet('其他管理器')
    ws_other.append(['节点类型', '功能描述', '所在文件'])

    other_managers = [
        # 环境管理器
        ['EnvironmentManagerNodeDefinition', '环境管理器主节点', 'environmentManagerNode.h'],
        ['PlayEnv_NodeType', '播放环境', 'environmentManagerNodeType.h'],
        ['PlayEnv_OverrideGlobalLight', '覆盖全局光照', 'environmentManagerNodeType.h'],
        ['PlayEnv_ForceRelitEnvProbe', '强制重新照明环境探针', 'environmentManagerNodeType.h'],
        ['PlayEnv_SetWeather', '设置天气', 'environmentManagerNodeType.h'],
        # 游戏管理器
        ['GameManagerNodeDefinition', '游戏管理器主节点', 'gameManagerNode.h'],
        ['TimeDilation_World', '世界时间膨胀', 'gameManagerNodeType.h'],
        ['TimeDilation_Player', '玩家时间膨胀', 'gameManagerNodeType.h'],
        ['TimeDilation_Entity', '实体时间膨胀', 'gameManagerNodeType.h'],
        ['ContentTokenManager_NodeType', '内容令牌管理器', 'gameManagerNodeType.h'],
        ['GameplayRestrictions_NodeType', '游戏限制', 'gameManagerNodeType.h'],
        ['SetTimer_NodeType', '设置计时器', 'gameManagerNodeType.h'],
        ['Rumble_NodeType', '震动', 'gameManagerNodeType.h'],
        # 事件管理器
        ['EventManagerNodeDefinition', '事件管理器节点', 'eventManagerNode.h'],
        # 特效管理器
        ['FXManagerNodeDefinition', '特效管理器主节点', 'fxManagerNode.h'],
        ['PlayFX_NodeType', '播放特效', 'fxManagerNodeType.h'],
        ['PreloadFX_NodeType', '预加载特效', 'fxManagerNodeType.h'],
        # 渲染特效管理器
        ['RenderFxManagerNodeDefinition', '渲染特效管理器主节点', 'renderFxManagerNode.h'],
        ['SetFadeInOut_NodeType', '设置淡入淡出', 'renderFxManagerNodeType.h'],
        ['SetDebugView_NodeType', '设置调试视图', 'renderFxManagerNodeType.h'],
        ['SetCyberspacePostFX_NodeType', '设置赛博空间后处理', 'renderFxManagerNodeType.h'],
        ['SetRenderLayer_NodeType', '设置渲染层', 'renderFxManagerNodeType.h'],
        # 物品管理器
        ['ItemManagerNodeDefinition', '物品管理器主节点', 'itemManagerNode.h'],
        ['AddRemoveItem_NodeType', '添加/移除物品', 'itemManagerNodeType.h'],
        ['DropItemFromSlot_NodeType', '从槽位丢弃物品', 'itemManagerNodeType.h'],
        ['SetItemTags_NodeType', '设置物品标签', 'itemManagerNodeType.h'],
        ['TransferItem_NodeType', '转移物品', 'itemManagerNodeType.h'],
        ['UseWeapon_NodeType', '使用武器', 'itemManagerNodeType.h'],
        ['InjectLoot_NodeType', '注入战利品', 'itemManagerNodeType.h'],
        # 交互对象管理器
        ['InteractiveObjectManagerNodeDefinition', '交互对象管理器主节点', 'interactiveObjectManagerNode.h'],
        ['SetInteractionState_NodeType', '设置交互状态', 'interactiveObjectManagerNodeType.h'],
        ['HackingManager_NodeType', '黑客管理器', 'interactiveObjectManagerNodeType.h'],
        ['DeviceManager_NodeType', '设备管理器', 'interactiveObjectManagerNodeType.h'],
        # 触发器管理器
        ['TriggerManagerNodeDefinition', '触发器管理器主节点', 'triggerManagerNode.h'],
        ['SetTriggerState_NodeType', '设置触发器状态', 'triggerManagerNodeType.h'],
        # 日志管理器
        ['JournalNodeDefinition', '日志节点主节点', 'journalNode.h'],
        ['JournalEntry_NodeType', '日志条目', 'journalNode.h'],
        ['JournalQuestEntry_NodeType', '任务日志条目', 'journalNode.h'],
        ['JournalTrackQuest_NodeType', '追踪任务', 'journalNode.h'],
        # 电话管理器
        ['PhoneManagerNodeDefinition', '电话管理器主节点', 'phoneManagerNode.h'],
        ['AddRemoveContact_NodeType', '添加/移除联系人', 'phoneManagerNodeType.h'],
        ['SetPhoneStatus_NodeType', '设置电话状态', 'phoneManagerNodeType.h'],
        ['CallContact_NodeType', '呼叫联系人', 'phoneManagerNodeType.h'],
        ['SendMessage_NodeType', '发送消息', 'phoneManagerNodeType.h'],
        # 场景管理器
        ['SceneManagerNodeDefinition', '场景管理器主节点', 'sceneManagerNode.h'],
        ['SetTier_NodeType', '设置Tier等级', 'sceneManagerNodeType.h'],
        ['PlayerLookAt_NodeType', '玩家注视', 'sceneManagerNodeType.h'],
        ['NPCLookAt_NodeType', 'NPC注视', 'sceneManagerNodeType.h'],
        ['SetFOV_NodeType', '设置视野', 'sceneManagerNodeType.h'],
        # 音频管理器
        ['AudioNodeDefinition', '音频节点主节点', 'audioNode.h'],
        ['AudioCharacterManagerNodeDefinition', '角色音频管理器', 'audioCharacterManagerNode.h'],
        ['AudioMixNodeType', '音频混合', 'audioNode.h'],
        ['AudioSwitchNodeType', '音频开关', 'audioSwitchNode.h'],
        # 行为管理器
        ['BehaviourManagerNodeDefinition', '行为管理器主节点', 'behaviourManagerNode.h'],
        ['JumpWorkspotAnim_NodeType', '跳转工作点动画', 'behaviourManagerNode.h'],
        ['StopWorkspot_NodeType', '停止工作点', 'behaviourManagerNode.h'],
        # 事实数据库管理器
        ['FactsDBManagerNodeDefinition', '事实数据库管理器主节点', 'factsDBManagerNode.h'],
        ['SetVar_NodeType', '设置变量', 'factsDBManagerNodeType.h'],
        # 其他
        ['MapPinManagerNodeDefinition', '地图标记管理器', 'mapPinManagerNode.h'],
        ['RewardManagerNodeDefinition', '奖励管理器主节点', 'rewardManagerNode.h'],
        ['GiveReward_NodeType', '给予奖励', 'rewardManagerNodeType.h'],
        ['SpawnManagerNodeDefinition', '生成管理器主节点', 'spawnManagerNode.h'],
        ['TimeManagerNodeDefinition', '时间管理器主节点', 'timeManagerNode.h'],
        ['VisionModesManagerNodeDefinition', '视觉模式管理器主节点', 'visionModesManagerNode.h'],
        ['VoicesetManagerNodeDefinition', '语音集管理器主节点', 'voicesetManagerNode.h'],
        ['RecordingNodeDefinition', '录制节点主节点', 'recordingNode.h'],
    ]

    for row in other_managers:
        ws_other.append(row)

    format_sheet(ws_other, header_font, header_fill, normal_font, border, alignment_center, alignment_left)

    # ==================== 10. 流程控制与杂项 ====================
    ws_misc = wb.create_sheet('流程控制与杂项')
    ws_misc.append(['节点类型', '功能描述', '所在文件'])

    misc_nodes = [
        # 流程控制
        ['FlowControlNodeDefinition', '流程控制节点', 'flowControlNode.h'],
        ['SwitchNodeDefinition', '开关节点', 'switchNode.h'],
        ['RandomizerNodeDefinition', '随机器节点', 'randomizerNode.h'],
        ['CheckpointNodeDefinition', '检查点节点', 'checkpointNode.h'],
        ['EmbeddedGraphNodeDefinition', '嵌入式图节点', 'embeddedGraphNode.h'],
        ['PhaseNodeDefinition', '阶段节点', 'phaseNode.h'],
        ['DeletionMarkerNodeDefinition', '删除标记节点', 'deletionMarkerNode.h'],
        # 多人游戏
        ['MultiplayerAIDirectorNodeDefinition', '多人游戏AI导演节点', 'multiplayerAIDirectorNode.h'],
        ['MultiplayerChoiceTokenNodeDefinition', '多人游戏选择令牌节点', 'multiplayerChoiceTokenNode.h'],
        ['MultiplayerJunctionDialogNodeDefinition', '多人游戏交汇对话节点', 'multiplayerJunctionDialogNode.h'],
        ['MultiplayerTeleportPuppetNodeDefinition', '多人游戏传送Puppet节点', 'multiplayerTeleportPuppetNode.h'],
        # 杂项
        ['BaseObjectNodeDefinition', '基础对象节点', 'baseObjectNode.h'],
        ['CutControlNodeDefinition', '剪辑控制节点', 'cutControlNode.h'],
        ['MinigameNodeDefinition', '小游戏节点', 'minigameNode.h'],
        ['PlaceholderNodeDefinition', '占位符节点', 'placeholderNode.h'],
        ['PuppeteerNodeDefinition', '操纵者节点', 'puppeteer.h'],
        ['PuppetAIManagerNodeDefinition', 'Puppet AI管理器节点', 'puppetAIManager.h'],
        ['PopulactionControllerNodeDefinition', '人口控制器节点', 'populationControllerNode.h'],
        ['InstancedCrowdControlNodeDefinition', '实例化人群控制节点', 'instancedCrowdControlNode.h'],
        ['TransformAnimatorNodeDefinition', '变换动画器节点', 'transformAnimatorNode.h'],
        ['TeleportVehicleNodeDefinition', '传送车辆节点', 'teleportVehicleNode.h'],
        ['WorkspotParamNodeDefinition', '工作点参数节点', 'workspotParamNode.h'],
    ]

    for row in misc_nodes:
        ws_misc.append(row)

    format_sheet(ws_misc, header_font, header_fill, normal_font, border, alignment_center, alignment_left)

    # ==================== 11. 统计汇总 ====================
    ws_summary = wb.create_sheet('统计汇总', 0)
    ws_summary.append(['分类', '节点数量', '说明'])

    summary_data = [
        ['核心节点', '12', '基础节点、开始/结束节点、I/O节点'],
        ['角色管理器', '28', '角色参数、战斗、视觉效果'],
        ['实体管理器', '16', '实体操作、附着、外观'],
        ['UI管理器', '48', '屏幕UI、弹窗、教程、通知'],
        ['车辆管理器', '30', '车辆控制、移动、战斗'],
        ['AI命令节点', '15', 'AI行为控制、移动、战斗'],
        ['逻辑节点', '4', '逻辑与、异或、Hub'],
        ['条件节点', '45', '对象、支付、属性、系统、时间条件'],
        ['其他管理器', '70', '环境、游戏、特效、物品等管理器'],
        ['流程控制与杂项', '22', '流程控制、多人游戏、杂项节点'],
        ['总计', '290+', '全部Quest节点类型'],
    ]

    for row in summary_data:
        ws_summary.append(row)

    # 格式化汇总表
    for row in ws_summary.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center
            cell.border = border

    for row in ws_summary.iter_rows(min_row=2):
        for idx, cell in enumerate(row):
            cell.font = normal_font
            cell.border = border
            if idx == 0:
                cell.alignment = alignment_left
            elif idx == 1:
                cell.alignment = alignment_center
            else:
                cell.alignment = alignment_left

    # 设置列宽
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 50

    # 保存文件
    output_file = r'E:\World\2077节点内容\quest_nodes.xlsx'
    wb.save(output_file)
    print(f'Excel 文件已创建: {output_file}')

def format_sheet(ws, header_font, header_fill, normal_font, border, alignment_center, alignment_left):
    """格式化工作表"""
    # 设置表头样式
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = border

    # 设置数据行样式
    for row in ws.iter_rows(min_row=2):
        for idx, cell in enumerate(row):
            cell.font = normal_font
            cell.border = border
            if idx == 0:  # 节点类型列
                cell.alignment = alignment_left
            elif idx == 1:  # 功能描述列
                cell.alignment = alignment_left
            else:  # 文件列
                cell.alignment = alignment_left

    # 设置列宽
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40

    # 冻结首行
    ws.freeze_panes = 'A2'

if __name__ == '__main__':
    create_excel()
